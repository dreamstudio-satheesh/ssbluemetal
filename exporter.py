"""Export reports to CSV, XLSX, and PDF summary table."""

import csv
import os
import tempfile

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors as reportlab_colors

from settings import get_company


# ── Colour helpers (mirrors printer.py) ──
NAVY = colors.HexColor("#1a237e")
LIGHT_NAVY = colors.HexColor("#e8eaf6")
BORDER = colors.HexColor("#cccccc")
WHITE = colors.white


def _get_font_name() -> str:
    """Try to register DejaVuSans; fall back to Helvetica."""
    from printer import FONT_NAME as printer_font
    return printer_font


FONT_NAME = _get_font_name()


# ═══════════════════════════════════════════════════════════════════
#  CSV
# ═══════════════════════════════════════════════════════════════════

def export_csv(rows: list[dict], column_headers: list[str], column_keys: list[str],
               output_path: str | None = None) -> str:
    """Write rows to CSV. Returns file path."""
    if output_path is None:
        output_path = os.path.join(tempfile.gettempdir(), "report.csv")

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(column_headers)
        for row in rows:
            writer.writerow([row.get(k, "") for k in column_keys])
    return output_path


# ═══════════════════════════════════════════════════════════════════
#  XLSX
# ═══════════════════════════════════════════════════════════════════

def export_xlsx(rows: list[dict], column_headers: list[str], column_keys: list[str],
                sheet_name: str = "Report",
                output_path: str | None = None) -> str:
    """Write rows to Excel .xlsx. Returns file path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if output_path is None:
        output_path = os.path.join(tempfile.gettempdir(), "report.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header style
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Write header
    for col_idx, header in enumerate(column_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    # Write data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(column_keys, 1):
            val = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left")

    # Auto-fit columns (rough)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════════════════
#  PDF Summary Table
# ═══════════════════════════════════════════════════════════════════

def export_pdf_table(rows: list[dict], column_headers: list[str], column_keys: list[str],
                     title: str = "Report",
                     output_path: str | None = None) -> str:
    """Write rows as a table in a PDF. Returns file path."""
    if output_path is None:
        output_path = os.path.join(tempfile.gettempdir(), "report.pdf")

    company = get_company()
    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        leftMargin=10*mm, rightMargin=10*mm,
        topMargin=12*mm, bottomMargin=12*mm,
    )

    styles = getSampleStyleSheet()
    s_title = ParagraphStyle("RptTitle", fontName=FONT_NAME, fontSize=16,
                             alignment=TA_CENTER, textColor=NAVY, spaceAfter=4)
    s_sub = ParagraphStyle("RptSub", fontName=FONT_NAME, fontSize=8,
                           alignment=TA_CENTER, textColor=colors.HexColor("#757575"),
                           spaceAfter=6*mm)
    s_hdr = ParagraphStyle("RptHdr", fontName=FONT_NAME, fontSize=7,
                           alignment=TA_CENTER, textColor=WHITE)
    s_cell = ParagraphStyle("RptCell", fontName=FONT_NAME, fontSize=7,
                            alignment=TA_CENTER)
    s_cell_left = ParagraphStyle("RptCellL", fontName=FONT_NAME, fontSize=7,
                                 alignment=TA_CENTER,  # will override per col
                                 )

    elements = []

    # Company header
    name = company.get("name", "").upper()
    if name:
        elements.append(Paragraph(name, s_title))
    gstin = company.get("gstin", "")
    if gstin:
        elements.append(Paragraph(f"GSTIN: {gstin}", s_sub))
    elements.append(Paragraph(title, ParagraphStyle(
        "Title2", fontName=FONT_NAME, fontSize=12,
        alignment=TA_CENTER, textColor=NAVY, spaceAfter=4*mm,
    )))

    # Build table data
    col_widths = [max(15*mm, 170*mm / len(column_headers))] * len(column_headers)
    data = [[Paragraph(h, s_hdr) for h in column_headers]]
    for row in rows:
        data.append([Paragraph(str(row.get(k, "")), s_cell) for k in column_keys])

    tbl = Table(data, colWidths=col_widths, repeatRows=1)

    # Row highlighting: header navy, total row light navy
    tbl_style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, -1), FONT_NAME),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]
    # Highlight last row if it looks like a total
    if len(data) > 2:
        last = data[-1]
        if any(kw in str(last[0]) for kw in ["Total", "Grand", "Summary"]):
            tbl_style_cmds.append(("BACKGROUND", (0, -1), (-1, -1), LIGHT_NAVY))
            tbl_style_cmds.append(("LINEABOVE", (0, -1), (-1, -1), 0.8, NAVY))

    tbl.setStyle(TableStyle(tbl_style_cmds))

    # If too many rows, use a smaller font
    if len(data) > 25:
        tbl_style_cmds.append(("FONTSIZE", (0, 1), (-1, -1), 6))

    elements.append(tbl)
    elements.append(Spacer(1, 4*mm))

    # Footer
    from reportlab.platypus import HRFlowable
    elements.append(HRFlowable(width="100%", thickness=0.3, color=BORDER, spaceAfter=2*mm))
    elements.append(Paragraph(
        "This is a computer-generated report.",
        ParagraphStyle("Footer", fontName=FONT_NAME, fontSize=7,
                       alignment=TA_CENTER, textColor=colors.HexColor("#757575")),
    ))

    doc.build(elements)
    return output_path
